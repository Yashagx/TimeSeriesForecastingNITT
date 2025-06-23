import os
import datetime
import tweepy
import requests
import pandas as pd
import torch
import demoji
import re
import numpy as np
from transformers import RobertaTokenizer, RobertaModel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === API KEYS ===
TWITTER_BEARER = 'AAAAAAAAAAAAAAAAAAAAAI3L2QEAAAAAo3LAhCkOPv2ouGm8qKV%2FA9WYM04%3DUt0sPDNtAOPdc4B11jeBf42YyVych65f5q2pNcSZEu100qk1ke'
NEWS_API_KEY = '1b6125d0311a4351994f2cb04c2ff887'
WEATHER_API_KEY = 'bd5e378503939ddaee76f12ad7a97608'

# === Load RoBERTa ===
tokenizer = RobertaTokenizer.from_pretrained("roberta-base")
model = RobertaModel.from_pretrained("roberta-base")

# === Clean Text ===
def clean_text(text):
    text = re.sub(r"http\S+", "", text)
    text = re.sub(r"@\w+", "", text)
    text = re.sub(r"#\w+", "", text)
    text = demoji.replace(text, "")
    text = re.sub(r"RT\s", "", text)
    return text.strip()

# === Get Embedding ===
def get_roberta_embedding(text):
    inputs = tokenizer(text, return_tensors="pt", truncation=True, padding=True, max_length=128)
    with torch.no_grad():
        outputs = model(**inputs)
    return outputs.last_hidden_state[:, 0, :].squeeze().tolist()

# === Fetch Tweets (English Only) ===
def fetch_tweets(query="COVID-19", max_results=10):
    client = tweepy.Client(bearer_token=TWITTER_BEARER)
    try:
        response = client.search_recent_tweets(query=f"{query} lang:en", max_results=max_results)
        return [tweet.text for tweet in response.data] if response.data else []
    except Exception as e:
        print("⚠️ Tweet fetch failed:", e)
        return []

# === Fetch News (English Only) ===
def fetch_news(query="COVID", country="us", page_size=10):
    url = "https://newsapi.org/v2/top-headlines"
    params = {"q": query, "apiKey": NEWS_API_KEY, "pageSize": page_size, "language": "en", "country": country}
    try:
        response = requests.get(url, params=params)
        return [article["title"] for article in response.json().get("articles", [])]
    except Exception as e:
        print("⚠️ News fetch failed:", e)
        return []

# === Clinical Data from OWID CSV ===
def fetch_clinical_from_csv(csv_path="owid-covid-data.csv", country="India"):
    try:
        df = pd.read_csv(csv_path)
        df = df[df["location"] == country].sort_values("date")
        latest = df.iloc[-1]
        return {
            "confirmed": round(latest.get("total_cases", 0), 2),
            "deaths": round(latest.get("total_deaths", 0), 2),
            "confirmed_per_million": round(latest.get("total_cases_per_million", 0), 2),
            "new_cases": round(latest.get("new_cases", 0), 2),
            "new_deaths": round(latest.get("new_deaths", 0), 2)
        }
    except Exception as e:
        print("⚠️ Clinical fetch failed:", e)
        return {
            "confirmed": None, "deaths": None,
            "confirmed_per_million": None,
            "new_cases": None, "new_deaths": None
        }

# === Climate Data ===
def fetch_climate(city="Chennai", country="IN"):
    try:
        url = f"http://api.openweathermap.org/data/2.5/weather?q={city},{country}&appid={WEATHER_API_KEY}&units=metric"
        data = requests.get(url).json()
        main = data.get("main", {})
        return {
            "temperature": round(main.get("temp", 0), 1),
            "feels_like": round(main.get("feels_like", 0), 1),
            "humidity": main.get("humidity"),
            "pressure": main.get("pressure"),
            "visibility": data.get("visibility", None),
            "wind_speed": data.get("wind", {}).get("speed", None)
        }
    except Exception as e:
        print("⚠️ Climate fetch failed:", e)
        return {"temperature": None, "humidity": None, "pressure": None}

# === Auto-fit Excel Columns ===
def auto_adjust_column_widths(path):
    wb = load_workbook(path)
    ws = wb.active
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 80)
    wb.save(path)

# === Main Pipeline ===
def run_daily_pipeline():
    print("🚀 Running multimodal pipeline...")
    today = datetime.date.today().strftime("%Y-%m-%d")
    folder = "multimodal_daily_data"
    os.makedirs(folder, exist_ok=True)

    tweets = fetch_tweets()
    news = fetch_news()
    texts = [clean_text(t) for t in tweets + news]

    clinical = fetch_clinical_from_csv()
    climate = fetch_climate()

    records = []
    embeddings = []

    for text in texts:
        emb = get_roberta_embedding(text)
        row = {
            "date": today,
            "text": text,
            **clinical,
            **climate
        }
        records.append(row)
        embeddings.append(emb)

    # Save metadata to Excel
    df_meta = pd.DataFrame(records)
    meta_path = f"{folder}/{today}_predictive_metadata.xlsx"
    df_meta.to_excel(meta_path, index=False, engine='openpyxl')
    auto_adjust_column_widths(meta_path)
    print(f"✅ Metadata saved: {meta_path}")

    # Save embeddings
    npy_path = f"{folder}/{today}_embeddings.npy"
    np.save(npy_path, np.array(embeddings))
    print(f"✅ Embeddings (.npy) saved: {npy_path}")

    # Optional: Save embeddings as CSV
    emb_csv_path = f"{folder}/{today}_embeddings.csv"
    pd.DataFrame(embeddings).to_csv(emb_csv_path, index=False)
    print(f"✅ Embeddings (.csv) saved: {emb_csv_path}")

# === Run Now ===
if __name__ == "__main__":
    run_daily_pipeline()
